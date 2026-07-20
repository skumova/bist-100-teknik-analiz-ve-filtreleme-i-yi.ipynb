"""
src/deep_value_data.py
======================

Derin-değer motoru (src/deep_value.py) için AĞ / VERİ ADAPTER katmanı.

Bu dosya açık internette (Colab) çalışır. İki tür veri sağlar:

1. OHLCV geçmişi  -> mevcut `src.data_loader.BistDataLoader` (tvdatafeed
   [rongardF] -> yfinance yedek) üzerinden günlük barlar.
2. Temel oranlar  -> ÖNCELİK: İş Yatırım public API (F/K, PD/DD, EV/EBITDA...);
   YEDEK: yfinance `.IS` .info (priceToBook, enterpriseToEbitda, trailingPE...).

Not: Bu ortamın (Claude oturumu) ağ politikası İş Yatırım/Yahoo'yu bloke eder;
bu yüzden fonksiyonlar burada değil, Colab'da çalıştırılmak üzere tasarlanmıştır.
Tümü try/except ile sarılıdır: bir kaynak başarısız olursa diğerine düşülür,
en kötü ihtimalle eksik alanlar None döner (motor eksiğe dayanıklıdır).
"""

from __future__ import annotations

import logging
from typing import Optional

import pandas as pd

logger = logging.getLogger("bist_bot.deep_value_data")
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
    logger.addHandler(handler)
logger.setLevel(logging.INFO)


# ---------------------------------------------------------------------- #
# Zaman aşımı yardımcısı (bir sembol askıda kalırsa taramayı kilitlemesin)
# ---------------------------------------------------------------------- #
_EXECUTOR = None


def _call_with_timeout(fn, timeout: float, *args, **kwargs):
    """fn'i ayrı bir thread'de çalıştırır; `timeout` sn içinde bitmezse
    TimeoutError yükseltir (askıdaki iş arka planda bırakılır, akış devam eder)."""
    global _EXECUTOR
    import concurrent.futures as _cf

    if _EXECUTOR is None:
        _EXECUTOR = _cf.ThreadPoolExecutor(max_workers=4)
    fut = _EXECUTOR.submit(fn, *args, **kwargs)
    return fut.result(timeout=timeout)


# ---------------------------------------------------------------------- #
# OHLCV
# ---------------------------------------------------------------------- #
def _normalize_ohlcv(raw: pd.DataFrame, n_bars: int) -> pd.DataFrame:
    """Farklı kaynaklardan gelen OHLCV'yi standart open/high/low/close/volume'a indirger."""
    df = raw.copy()
    df.columns = [str(c).lower() for c in df.columns]
    if "close" not in df.columns and "adj close" in df.columns:
        df = df.rename(columns={"adj close": "close"})
    missing = [c for c in ("open", "high", "low", "close", "volume") if c not in df.columns]
    if missing:
        raise RuntimeError(f"eksik OHLCV kolonları: {missing} (mevcut: {list(df.columns)})")
    df = df[["open", "high", "low", "close", "volume"]]
    if getattr(df.index, "tz", None) is not None:
        df.index = df.index.tz_localize(None)
    df.index.name = "datetime"
    return df.sort_index().tail(n_bars)


def _period_for(n_bars: int) -> str:
    """n_bars işlem gününü kapsayacak borsapy/yfinance period etiketi."""
    if n_bars <= 480:
        return "2y"
    if n_bars <= 1200:
        return "5y"
    return "10y"


def _fetch_borsapy_daily(symbol: str, n_bars: int) -> pd.DataFrame:
    """borsapy ile günlük OHLCV — BIST için EN TEMİZ/SAĞLIKLI kaynak (TradingView backend)."""
    import borsapy

    raw = borsapy.Ticker(symbol.upper()).history(
        period=_period_for(n_bars), interval="1d", auto_adjust=False
    )
    if raw is None or len(raw) == 0:
        raise RuntimeError(f"borsapy boş veri döndürdü: {symbol}")
    return _normalize_ohlcv(raw, n_bars)


def _fetch_yf_daily(symbol: str, n_bars: int) -> pd.DataFrame:
    """yfinance ile günlük OHLCV (yedek). Not: BIST'te bazı sembollerde eksik/hatalı olabilir."""
    import yfinance as yf

    sym = symbol.upper()
    if not sym.endswith(".IS"):
        sym += ".IS"
    raw = yf.download(sym, period=_period_for(n_bars), interval="1d", progress=False,
                      auto_adjust=False, multi_level_index=False)
    if raw is None or raw.empty:
        raise RuntimeError(f"yfinance boş veri döndürdü: {sym}")
    return _normalize_ohlcv(raw, n_bars)


# Her "source" için denenecek kaynak zinciri (ilki asıl, kalanı yedek)
_SOURCE_CHAINS = {
    "borsapy": ["borsapy", "yfinance"],
    "tvdatafeed": ["tvdatafeed", "yfinance"],
    "yfinance": ["yfinance"],
}


def load_daily_ohlcv(symbol: str, n_bars: int = 600, loader=None, source: str = "borsapy",
                     timeout: float = 25.0, return_source: bool = False):
    """Günlük OHLCV döndürür. BIST için önerilen kaynak sırası borsapy → yfinance.

    source="borsapy" (VARSAYILAN): borsapy (TradingView backend) — BIST'te en
        temiz/sağlıklı veri. Başarısız olur / `timeout` sn içinde bitmezse O SEMBOL
        için yfinance'a düşülür.
    source="tvdatafeed": TradingView (rongardF fork) `BistDataLoader` üzerinden;
        yedeği yfinance.
    source="yfinance": yalnızca yfinance.

    return_source=True ise (df, kullanılan_kaynak) döner. Her fetch bir thread +
    zaman aşımıyla sarılıdır; bir sembol askıda kalırsa atlanır, tarama kilitlenmez.
    """
    chain = _SOURCE_CHAINS.get(source, [source])
    last_exc = None
    for src in chain:
        try:
            if src == "tvdatafeed":
                if loader is None:
                    from src.data_loader import BistDataLoader

                    loader = BistDataLoader(interval="1d")
                df = _call_with_timeout(loader.get_history, timeout, symbol, n_bars=n_bars)
            elif src == "borsapy":
                df = _call_with_timeout(_fetch_borsapy_daily, timeout, symbol, n_bars)
            else:
                df = _call_with_timeout(_fetch_yf_daily, timeout, symbol, n_bars)
            if df is not None and len(df) >= 60:
                return (df, src) if return_source else df
            last_exc = RuntimeError(f"{src}: yetersiz bar")
        except Exception as exc:  # noqa: BLE001  (TimeoutError dahil)
            last_exc = exc
            logger.debug("[%s] %s başarısız (%s), sıradaki kaynağa geçiliyor.", symbol, src, exc)
    raise RuntimeError(f"{symbol}: hiçbir kaynaktan veri alınamadı ({last_exc})")


def average_tl_volume(df: pd.DataFrame, window: int = 20) -> Optional[float]:
    """Son `window` bardaki ortalama TL bazlı işlem hacmi (likidite tuzağı için)."""
    if df is None or df.empty:
        return None
    tl = (df["close"] * df["volume"]).tail(window)
    return float(tl.mean()) if len(tl) else None


# ---------------------------------------------------------------------- #
# Temel oranlar - İş Yatırım (öncelik)
# ---------------------------------------------------------------------- #
_ISYATIRIM_URL = (
    "https://www.isyatirim.com.tr/_layouts/15/IsYatirim.Website/Common/"
    "Data.aspx/StockAnalysisRatios?hisse={sym}"
)


def fetch_ratios_isyatirim(symbol: str, timeout: int = 15) -> dict:
    """İş Yatırım'dan temel oranları çeker (best-effort).

    İş Yatırım'ın public JSON uçları zamanla değişebilir; bu yüzden fonksiyon
    başarısız olursa BOŞ dict döner ve çağıran taraf yfinance yedeğine düşer.
    Beklenen çıktı anahtarları motorun (deep_value) beklediği isimlerdir.
    """
    import requests

    try:
        url = _ISYATIRIM_URL.format(sym=symbol.upper())
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        payload = r.json()
        rows = payload.get("value") or payload.get("d") or []
        if not rows:
            return {}
        rec = rows[0] if isinstance(rows, list) else rows
        out = {
            "pe_ratio": rec.get("FK") or rec.get("FiyatKazanc"),
            "pb_ratio": rec.get("PDDD") or rec.get("PiyasaDegeriDefterDegeri"),
            "ev_ebitda": rec.get("FDFAVOK") or rec.get("EVEBITDA"),
            "ev_sales": rec.get("FDSatis") or rec.get("EVSatis"),
        }
        return {k: v for k, v in out.items() if v is not None}
    except Exception as exc:  # noqa: BLE001
        logger.debug("[%s] İş Yatırım oranları alınamadı: %s", symbol, exc)
        return {}


# ---------------------------------------------------------------------- #
# Temel oranlar - yfinance yedek (portatif, Colab'da güvenilir çalışır)
# ---------------------------------------------------------------------- #
def fetch_ratios_yfinance(symbol: str) -> dict:
    """yfinance `.IS` .info'dan temel oranlar (yedek kaynak).

    Motorun beklediği anahtar isimlerine map'ler. Eksik alanlar döndürülmez.
    """
    try:
        import yfinance as yf

        sym = symbol.upper()
        if not sym.endswith(".IS"):
            sym = f"{sym}.IS"
        info = yf.Ticker(sym).info or {}
    except Exception as exc:  # noqa: BLE001
        logger.debug("[%s] yfinance info alınamadı: %s", symbol, exc)
        return {}

    out = {
        "pe_ratio": info.get("trailingPE"),
        "pb_ratio": info.get("priceToBook"),
        "ev_ebitda": info.get("enterpriseToEbitda"),
        "ev_sales": info.get("enterpriseToRevenue"),
        "roe": info.get("returnOnEquity"),
        "net_debt_to_equity": (info.get("debtToEquity") / 100.0) if info.get("debtToEquity") else None,
        "net_income": info.get("netIncomeToCommon"),
        "owner_earnings": info.get("freeCashflow"),  # yaklaşık: serbest nakit akışı
        "sector": info.get("sector"),
    }
    return {k: v for k, v in out.items() if v is not None}


def fetch_fundamentals(symbol: str, prefer_isyatirim: bool = True) -> dict:
    """İş Yatırım + yfinance oranlarını birleştirir (İş Yatırım öncelikli)."""
    yf_data = fetch_ratios_yfinance(symbol)
    if not prefer_isyatirim:
        return yf_data
    isy = fetch_ratios_isyatirim(symbol)
    merged = dict(yf_data)
    merged.update({k: v for k, v in isy.items() if v is not None})  # İş Yatırım üstün gelir
    return merged


# ---------------------------------------------------------------------- #
# Universe tarama (uçtan uca)
# ---------------------------------------------------------------------- #
def screen_universe(
    symbols: list[str],
    n_bars: int = 600,
    fib_lookback: int = 180,
    prefer_isyatirim: bool = True,
    with_ladder_top_n: int = 15,
    tech_prefilter: bool = True,
    prefilter_margin: float = 10.0,
    source: str = "borsapy",
    progress_every: int = 25,
    fetch_timeout: float = 12.0,
):
    """Sembol listesini uçtan uca tarar: veri çek -> skorla -> raporla.

    ANA belirleyici teknik ucuzluk olduğu için önce teknik hesaplanır; teknik
    kapının (TECH_GATE) belirgin altında kalan hisseler için pahalı temel veri
    çekilmez (`tech_prefilter=True`). Böylece 100 hissede gereksiz API isteği
    yapılmaz — sadece "aşırı ucuz" adaylara temel + banker ekstra puanı işlenir.

    Dönüş: (rapor_df, results) — rapor_df sıralı özet tablo, results ise
    her sembol için tam DeepValueResult listesi (Fibonacci merdivenleri dahil).
    """
    import sys as _sys, os as _os, warnings as _w, logging as _lg, contextlib as _cl
    from src import deep_value as dv

    # Gürültülü çıktıyı sustur — yüzlerce sembolde tvdatafeed/yfinance'ın bastığı
    # binlerce log/uyarı satırı Colab sekmesini DONDURUR. Bunu kökten engelle.
    _w.filterwarnings("ignore")
    for _n in ("bist_bot", "bist_bot.deep_value_data", "bist_bot.data_loader",
               "tvDatafeed", "tvDatafeed.main", "websocket", "yfinance", "urllib3", "peewee"):
        _lg.getLogger(_n).setLevel(_lg.CRITICAL)
    _real_out = _sys.stdout           # ilerleme yazıları için (bastırılmayan) gerçek çıktı
    _devnull = open(_os.devnull, "w")

    loader = None
    if source == "tvdatafeed":
        from src.data_loader import BistDataLoader
        loader = BistDataLoader(interval="1d")

    results = []
    failed = 0
    primary_miss = 0                 # birincil kaynağın (borsapy/tvdatafeed) üst üste ıskalaması
    effective_source = source
    gate = dv.TECH_GATE - prefilter_margin
    total = len(symbols)
    for i, sym in enumerate(symbols, 1):
        try:
            # Sembol başına TÜM stdout/stderr'i /dev/null'a yönlendir (çıktı seli = donma)
            with _cl.redirect_stdout(_devnull), _cl.redirect_stderr(_devnull):
                df, used = load_daily_ohlcv(sym, n_bars=n_bars, loader=loader,
                                            source=effective_source, timeout=fetch_timeout,
                                            return_source=True)
                if df is None or len(df) < 60:
                    raise RuntimeError("yetersiz OHLCV")
                # Teknik ön-eleme: kapıyı geçemeyecek kadar pahalıysa temel veri çekme
                if tech_prefilter and dv.technical_cheapness(df).total < gate:
                    ratios = {}
                else:
                    ratios = fetch_fundamentals(sym, prefer_isyatirim=prefer_isyatirim)
                res = dv.composite_score(
                    sym, df, ratios,
                    sector=ratios.get("sector"),
                    avg_tl_volume=average_tl_volume(df),
                )
            results.append((res, df, ratios))
            # Birincil kaynak ıskaladı mı? (veri geldi ama yedekten)
            primary_miss = 0 if (effective_source == "yfinance" or used == effective_source) else primary_miss + 1
        except Exception:  # noqa: BLE001  (veri yok / zaman aşımı / hesap hatası)
            failed += 1
            primary_miss += (effective_source != "yfinance")
        # Birincil kaynak (borsapy/tvdatafeed) sürekli ıskalıyorsa kalanları doğrudan
        # yfinance'a çevir — her sembolde timeout beklemenin ağır bedelini önle.
        if effective_source != "yfinance" and primary_miss >= 6:
            _real_out.write(f"  ⚠ '{effective_source}' bu ortamda çalışmıyor; kalan semboller "
                            f"yfinance ile çekiliyor.\n")
            _real_out.flush()
            effective_source, loader = "yfinance", None
        # İlerleme: yalnızca bunu göster (gerçek çıktıya, bastırılmadan)
        if progress_every and (i % progress_every == 0 or i == total):
            _real_out.write(f"  … {i}/{total} tarandı | {len(results)} geçerli, {failed} atlandı "
                            f"[{effective_source}]\n")
            _real_out.flush()

    ranked = sorted((r[0] for r in results), key=lambda x: x.final_score, reverse=True)
    # En iyi N adaya Fibonacci merdiveni ekle
    df_by_sym = {r[0].symbol: r[1] for r in results}
    ratios_by_sym = {r[0].symbol: r[2] for r in results}
    for res in ranked[:with_ladder_top_n]:
        d = df_by_sym[res.symbol]
        dcf = None
        r = ratios_by_sym[res.symbol]
        if r.get("dcf_intrinsic_value"):
            dcf = dv._num(r.get("dcf_intrinsic_value"))
        res.ladder = dv.fibonacci_ladder(d, lookback=fib_lookback, dcf_target=dcf)

    report = dv.build_report(ranked)
    return report, ranked


# ---------------------------------------------------------------------- #
# Detaylı Excel çıktısı
# ---------------------------------------------------------------------- #
def _ladder_rows(results) -> "pd.DataFrame":
    """Fibonacci merdivenlerini uzun-format (her kademe bir satır) tabloya çevirir."""
    rows = []
    for res in results:
        if res.ladder is None:
            continue
        L = res.ladder
        for i, r in enumerate(L.rungs, 1):
            rows.append({
                "symbol": res.symbol,
                "final_score": res.final_score,
                "kademe": i,
                "fib_orani": r["ratio"],
                "fiyat": r["price"],
                "agirlik_%": r["weight_pct"],
                "aciklama": r["note"],
                "swing_high": L.swing_high,
                "swing_low": L.swing_low,
                "guncel_fiyat": L.current_price,
                "hard_stop": L.hard_stop,
                "dcf_hedef": L.dcf_target,
                "beklenen_getiri_%": L.expected_upside_pct,
            })
    return pd.DataFrame(rows)


def _detail_rows(results) -> "pd.DataFrame":
    """Her hisse için tüm alt bileşenleri (teknik/banker/temel ham + skor) tek tabloda."""
    rows = []
    for res in results:
        t, b, f = res.technical, res.banker, res.fundamental
        rows.append({
            "symbol": res.symbol,
            "sector": res.sector,
            "final_score": res.final_score,
            "asiri_ucuz": res.qualifies,
            "cekirdek": res.core_score,
            "tech_ucuzluk": t.total,
            "banker": b.total,
            "temel_skor": f.total,
            "temel_carpan": res.fund_bonus,
            "trap_carpan": res.trap.multiplier,
            "diskalifiye": res.trap.disqualified,
            # Teknik ham
            "rsi_d": round(t.raw.get("rsi_d"), 1) if t.raw.get("rsi_d") is not None else None,
            "rsi_w": round(t.raw["rsi_w"], 1) if t.raw.get("rsi_w") is not None else None,
            "pos_52w": round(t.raw["pos_52w"], 3) if t.raw.get("pos_52w") is not None else None,
            "drawdown": round(t.raw["drawdown"], 3) if t.raw.get("drawdown") is not None else None,
            "bb_percent_b": round(t.raw["bb_percent_b"], 3) if t.raw.get("bb_percent_b") is not None else None,
            "ema200_dev": round(t.raw["ema200_dev"], 3) if t.raw.get("ema200_dev") is not None else None,
            "williams_r": round(t.raw["williams_r"], 1) if t.raw.get("williams_r") is not None else None,
            "high_52w": t.raw.get("high_52w"),
            "low_52w": t.raw.get("low_52w"),
            "dip_confirm": t.dip_confirm,
            # Banker ham
            "cmf": round(b.raw["cmf"], 3),
            "mfi": round(b.raw["mfi"], 1),
            "ad_slope": round(b.raw["ad_slope"], 2),
            "obv_slope": round(b.raw["obv_slope"], 2),
            "birikim_diverjans": b.accumulation,
            # Temel skor bileşenleri
            "value": f.value,
            "quality": f.quality,
            # Value-trap
            "trap_flags": "; ".join(res.trap.flags) if res.trap.flags else "",
        })
    return pd.DataFrame(rows).sort_values("final_score", ascending=False).reset_index(drop=True)


def export_to_excel(report, results, path: str = "bist_derin_deger_tarama.xlsx", capital: float = 100_000.0) -> str:
    """Taramayı çok-sayfalı, biçimlendirilmiş detaylı bir Excel dosyasına yazar.

    Sayfalar:
      - Ozet        : sıralı özet tablo (build_report çıktısı)
      - Adaylar     : teknik kapıyı geçen, tuzaksız hisseler
      - Tuzaklar    : value-trap bayraklı / diskalifiye edilenler
      - Detay       : tüm alt bileşenler (teknik/banker/temel ham değerler)
      - Alim_Plani  : Fibonacci 3-kademe merdivenleri (uzun format) + TL/lot dağılımı
    """
    from src import deep_value as dv

    detail = _detail_rows(results)
    adaylar = detail[detail["asiri_ucuz"] & (~detail["diskalifiye"])].reset_index(drop=True)
    tuzaklar = detail[(detail["trap_flags"] != "") | (detail["diskalifiye"])].reset_index(drop=True)
    ladder = _ladder_rows(results)
    if not ladder.empty:
        ladder["kademe_TL"] = (capital * ladder["agirlik_%"] / 100).round(0)
        ladder["yakl_lot"] = (ladder["kademe_TL"] / ladder["fiyat"]).astype(int)

    try:
        import openpyxl  # noqa
        engine = "openpyxl"
    except ImportError:
        engine = None  # pandas varsayılanına bırak

    with pd.ExcelWriter(path, engine=engine) as xl:
        report.to_excel(xl, sheet_name="Ozet", index=False)
        adaylar.to_excel(xl, sheet_name="Adaylar", index=False)
        tuzaklar.to_excel(xl, sheet_name="Tuzaklar", index=False)
        detail.to_excel(xl, sheet_name="Detay", index=False)
        if not ladder.empty:
            ladder.to_excel(xl, sheet_name="Alim_Plani", index=False)
        _format_workbook(xl)

    logger.info("Excel yazıldı: %s (%d hisse, %d aday)", path, len(detail), len(adaylar))
    return path


def _format_workbook(xl) -> None:
    """Kolon genişliği + başlık + koşullu renk (openpyxl varsa)."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
    except Exception:  # noqa: BLE001
        return
    wb = xl.book
    header_fill = PatternFill("solid", fgColor="1F2D3D")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        # Başlık satırı biçimi
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        # Kolon genişlikleri
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 9), 42)
        # final_score kolonuna renk skalası (yeşil=yüksek)
        header = {c.value: c.column_letter for c in ws[1]}
        if "final_score" in header and ws.max_row > 1:
            col = header["final_score"]
            rng = f"{col}2:{col}{ws.max_row}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=40, mid_color="FFEB84",
                end_type="num", end_value=80, end_color="63BE7B"))
